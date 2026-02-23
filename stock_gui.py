import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import threading
import queue
from datetime import datetime, timedelta
import json
import os

import pandas as pd
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.dates as mdates

from stock_volume_predictor import StockTradingSystem

# Fixed filenames so updates always append to the same files
STOCK_DATA_FILE   = "stock_data.xlsx"
PREDICTIONS_FILE  = "stock_predictions.xlsx"


class StockPriceGUI:
    """
    Stock price predictor GUI with:
    - Tab per stock showing actual vs predicted close price chart
    - Excel update functions that preserve historical data
    """

    def __init__(self, root):
        self.root = root
        self.root.title("Stock Price Predictor — Yahoo Finance")
        self.root.geometry("1100x750")

        self.stocks = {}        # {symbol: {system, prediction, raw_df, pred_history, status}}
        self.message_queue = queue.Queue()

        self.create_widgets()
        self.process_queue()

    # ------------------------------------------------------------------ #
    #  UI CONSTRUCTION                                                     #
    # ------------------------------------------------------------------ #

    def create_widgets(self):
        root_nb = ttk.Notebook(self.root)
        root_nb.pack(fill="both", expand=True, padx=6, pady=6)

        # ── Tab 1: Stock Manager ──────────────────────────────────── #
        mgr_frame = ttk.Frame(root_nb)
        root_nb.add(mgr_frame, text="  Stock Manager  ")
        mgr_frame.columnconfigure(0, weight=1)
        mgr_frame.rowconfigure(2, weight=1)

        # Controls
        ctrl = ttk.LabelFrame(mgr_frame, text="Add Stock", padding="6")
        ctrl.grid(row=0, column=0, sticky="ew", padx=8, pady=6)

        ttk.Label(ctrl, text="Symbol:").grid(row=0, column=0, padx=4)
        self.symbol_var = tk.StringVar()
        ttk.Entry(ctrl, textvariable=self.symbol_var, width=10).grid(row=0, column=1, padx=4)

        ttk.Label(ctrl, text="Lookback:").grid(row=0, column=2, padx=4)
        self.lookback_var = tk.IntVar(value=10)
        ttk.Spinbox(ctrl, from_=5, to=30, textvariable=self.lookback_var, width=6).grid(row=0, column=3, padx=4)

        ttk.Label(ctrl, text="Epochs:").grid(row=0, column=4, padx=4)
        self.epochs_var = tk.IntVar(value=200)
        ttk.Spinbox(ctrl, from_=50, to=500, textvariable=self.epochs_var, width=7).grid(row=0, column=5, padx=4)

        ttk.Button(ctrl, text="Add & Train",               command=self.add_stock).grid(row=0, column=6, padx=5)
        ttk.Button(ctrl, text="Quick Add (SPY/AAPL/MSFT)", command=self.quick_add).grid(row=0, column=7, padx=5)

        # Stock table
        tbl = ttk.LabelFrame(mgr_frame, text="Tracked Stocks", padding="6")
        tbl.grid(row=2, column=0, sticky="nsew", padx=8, pady=4)
        tbl.columnconfigure(0, weight=1)
        tbl.rowconfigure(0, weight=1)

        cols = ("Symbol", "Current Price", "Sentiment", "Confidence", "Signal", "Status")
        self.tree = ttk.Treeview(tbl, columns=cols, show="headings", height=10)
        for col, w in zip(cols, [90, 110, 120, 100, 130, 150]):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, anchor="center")
        self.tree.tag_configure("ready",    background="#e8f5e9")
        self.tree.tag_configure("training", background="#fff9c4")
        self.tree.tag_configure("error",    background="#ffebee")
        vsb = ttk.Scrollbar(tbl, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        # Action buttons
        btn = ttk.Frame(mgr_frame)
        btn.grid(row=3, column=0, sticky="w", padx=8, pady=4)
        for text, cmd in [
            ("Predict Selected",    self.predict_selected),
            ("Predict All",         self.predict_all),
            ("Update Selected",     self.update_selected),
            ("Update All",          self.update_all),
            ("Remove Selected",     self.remove_selected),
            ("Export Stock Data",   self.export_stock_data),
            ("Update Stock Data",   self.update_stock_data),
            ("Export Predictions",  self.export_predictions),
            ("Update Predictions",  self.update_predictions),
        ]:
            ttk.Button(btn, text=text, command=cmd).pack(side="left", padx=2)

        # Log
        log_f = ttk.LabelFrame(mgr_frame, text="Log", padding="4")
        log_f.grid(row=4, column=0, sticky="ew", padx=8, pady=4)
        log_f.columnconfigure(0, weight=1)
        self.log_text = scrolledtext.ScrolledText(
            log_f, height=6, wrap=tk.WORD, font=("Courier", 9)
        )
        self.log_text.grid(row=0, column=0, sticky="ew")

        # Status bar
        self.status_var = tk.StringVar(value="Ready — add stocks to begin")
        ttk.Label(mgr_frame, textvariable=self.status_var, relief="sunken", anchor="w"
                  ).grid(row=5, column=0, sticky="ew", padx=8)

        # ── Tab 2: Charts notebook ────────────────────────────────── #
        chart_outer = ttk.Frame(root_nb)
        root_nb.add(chart_outer, text="  Charts  ")
        chart_outer.columnconfigure(0, weight=1)
        chart_outer.rowconfigure(1, weight=1)

        chart_ctrl = ttk.Frame(chart_outer)
        chart_ctrl.grid(row=0, column=0, sticky="ew", padx=8, pady=6)
        ttk.Button(chart_ctrl, text="Refresh Charts", command=self.refresh_all_charts).pack(side="left", padx=4)

        self.chart_nb = ttk.Notebook(chart_outer)
        self.chart_nb.grid(row=1, column=0, sticky="nsew", padx=8, pady=4)

        self._chart_tabs   = {}   # symbol → frame
        self._chart_canvas = {}   # symbol → FigureCanvasTkAgg

    # ------------------------------------------------------------------ #
    #  LOGGING                                                             #
    # ------------------------------------------------------------------ #

    def log(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{ts}] {msg}\n")
        self.log_text.see(tk.END)
        print(msg)

    # ------------------------------------------------------------------ #
    #  TREE HELPERS                                                        #
    # ------------------------------------------------------------------ #

    def _tree_has(self, iid):
        try:
            self.tree.item(iid)
            return True
        except tk.TclError:
            return False

    def _update_tree_row(self, symbol):
        data = self.stocks.get(symbol, {})
        pred = data.get("prediction")
        status = data.get("status", "—")
        sig_bg = sig_fg = sen_bg = sen_fg = None
        if pred:
            current   = f"${pred['current_price']:.2f}"
            sentiment = pred.get("sentiment_analysis", {}).get("sentiment", "—")
            conf      = f"{pred['confidence'] * 100:.0f}%"
            signal    = pred.get("recommendation", "—").replace("_", " ")

            # Build unique tag names per symbol so colours don't bleed across rows
            sig_bg, sig_fg   = self._signal_colours(signal)
            sen_bg, sen_fg   = self._sentiment_colours(sentiment)
            sig_tag = f"sig_{symbol}"
            sen_tag = f"sen_{symbol}"
            if sig_bg and sig_fg:
                self.tree.tag_configure(sig_tag, background=sig_bg, foreground=sig_fg)
            if sen_bg and sen_fg:
                self.tree.tag_configure(sen_tag, background=sen_bg, foreground=sen_fg)

            tag = "ready"
        else:
            current = sentiment = conf = signal = "—"
            sig_tag = sen_tag = None
            tag = "error" if "Error" in status else "training"

        vals = (symbol, current, sentiment, conf, signal, status)
        if self._tree_has(symbol):
            self.tree.item(symbol, values=vals, tags=(tag,))
        else:
            self.tree.insert("", "end", iid=symbol, values=vals, tags=(tag,))

        # Colour individual cells via a Canvas overlay drawn on top of the Treeview.
        # tkinter's Treeview doesn't support per-cell colour natively, so we use
        # a tag per row and recolour the whole row then redraw named-column text.
        # The simplest compatible approach: use per-row tags with the signal colour
        # for the Signal column by making each symbol's row its own tag.
        if pred and (sig_bg or sen_bg):
            row_tag = f"row_{symbol}"
            # We can only set one background per row with Treeview tags.
            # Prefer signal colour for the row; sentiment shown via text prefix.
            if sig_bg and sig_fg:
                self.tree.tag_configure(row_tag, background=sig_bg, foreground=sig_fg)
                self.tree.item(symbol, tags=(row_tag,))
            # Update signal text to include sentiment indicator
            sen_icon = ""
            if sentiment in ("Very Bullish", "Bullish"):
                sen_icon = "▲ "
            elif sentiment in ("Very Bearish", "Bearish"):
                sen_icon = "▼ "
            else:
                sen_icon = "— "
            # Rebuild values with sentiment icon
            vals = (symbol, current, sen_icon + sentiment, conf, signal, status)
            self.tree.item(symbol, values=vals)

    def _selected_symbols(self):
        return list(self.tree.selection())

    # ------------------------------------------------------------------ #
    #  CHART MANAGEMENT                                                    #
    # ------------------------------------------------------------------ #

    # ------------------------------------------------------------------ #
    #  GRADIENT COLOUR HELPERS                                             #
    # ------------------------------------------------------------------ #

    # Signal order: STRONG SELL → SELL → HOLD → BUY → STRONG BUY
    # Colour gradient: red (#d32f2f) → orange → grey → light-green → green (#2e7d32)
    _SIGNAL_MAP = {
        "STRONG SELL": 0,
        "SELL":        1,
        "HOLD":        2,
        "BUY":         3,
        "STRONG BUY":  4,
    }
    _SIGNAL_COLOURS = ["#d32f2f", "#e57373", "#bdbdbd", "#81c784", "#2e7d32"]
    _SIGNAL_FG      = ["#ffffff", "#000000", "#000000", "#000000", "#ffffff"]

    # Sentiment: Very Bearish → Bearish → Neutral → Bullish → Very Bullish
    _SENTIMENT_MAP = {
        "Very Bearish": 0,
        "Bearish":      1,
        "Neutral":      2,
        "Bullish":      3,
        "Very Bullish": 4,
    }
    _SENTIMENT_COLOURS = ["#d32f2f", "#e57373", "#bdbdbd", "#81c784", "#2e7d32"]
    _SENTIMENT_FG      = ["#ffffff", "#000000", "#000000", "#000000", "#ffffff"]

    def _signal_colours(self, signal_text):
        """Return (bg, fg) for a signal string."""
        key = signal_text.strip().upper().replace("_", " ")
        idx = self._SIGNAL_MAP.get(key, -1)
        if idx == -1:
            return None, None
        return self._SIGNAL_COLOURS[idx], self._SIGNAL_FG[idx]

    def _sentiment_colours(self, sentiment_text):
        """Return (bg, fg) for a sentiment string."""
        key = sentiment_text.strip().title()
        idx = self._SENTIMENT_MAP.get(key, -1)
        if idx == -1:
            return None, None
        return self._SENTIMENT_COLOURS[idx], self._SENTIMENT_FG[idx]

    # ------------------------------------------------------------------ #
    #  CHART MANAGEMENT                                                    #
    # ------------------------------------------------------------------ #

    def _ensure_chart_tab(self, symbol):
        """Create chart tab for symbol if it doesn't exist."""
        if symbol in self._chart_tabs:
            return
        frame = ttk.Frame(self.chart_nb)
        self.chart_nb.add(frame, text=f"  {symbol}  ")
        self._chart_tabs[symbol] = frame

    def _draw_chart(self, symbol):
        """Draw actual vs predicted chart with hover crosshair tooltip."""
        data = self.stocks.get(symbol)
        if not data:
            return

        self._ensure_chart_tab(symbol)
        frame = self._chart_tabs[symbol]

        # Clear previous widgets
        if symbol in self._chart_canvas:
            try:
                self._chart_canvas[symbol].get_tk_widget().destroy()
            except Exception:
                pass
        for widget in frame.winfo_children():
            widget.destroy()

        df   = data.get("raw_df")
        pred = data.get("prediction")
        pred_history = data.get("pred_history", [])

        fig, (ax_main, ax_pred) = plt.subplots(
            2, 1, figsize=(11, 7),
            gridspec_kw={"height_ratios": [3, 2]},
            sharex=False
        )
        fig.subplots_adjust(hspace=0.45)

        # ── TOP: Actual close ──────────────────────────────────────── #
        ax_main.grid(color="#e0e0e0", linewidth=0.7, linestyle="--")

        actual_dates  = []
        actual_closes = []

        if df is not None and len(df) > 0:
            df_plot = df.copy()
            if hasattr(df_plot.index, "tz") and df_plot.index.tz is not None:
                df_plot.index = df_plot.index.tz_localize(None)
            actual_dates  = list(df_plot.index)
            actual_closes = list(df_plot["close"])
            ax_main.plot(actual_dates, actual_closes,
                         color="#1565c0", linewidth=1.5, label="Actual Close", zorder=3)

        ax_main.set_title(f"{symbol} — Actual Close Price", fontsize=10, pad=8)
        ax_main.set_ylabel("Price (USD)", fontsize=9)
        ax_main.xaxis.set_major_formatter(mdates.DateFormatter("%b %y"))
        ax_main.xaxis.set_major_locator(mdates.MonthLocator())
        plt.setp(ax_main.xaxis.get_majorticklabels(), rotation=30, ha="right")
        ax_main.legend(fontsize=8)

        # Crosshair elements for top chart
        vline_main = ax_main.axvline(x=actual_dates[0] if actual_dates else datetime.now(),
                                     color="gray", linewidth=0.8, linestyle=":", visible=False)
        hline_main = ax_main.axhline(y=0, color="gray", linewidth=0.8, linestyle=":", visible=False)
        dot_main,  = ax_main.plot([], [], "o", color="#1565c0", markersize=6, zorder=6)
        tooltip_main = ax_main.annotate(
            "", xy=(0, 0), xytext=(12, 12), textcoords="offset points",
            bbox=dict(boxstyle="round,pad=0.4", fc="white", ec="#555", alpha=0.9),
            fontsize=8, visible=False
        )

        # ── BOTTOM: Prediction scenarios ───────────────────────────── #
        ax_pred.grid(color="#e0e0e0", linewidth=0.7, linestyle="--")

        all_pred_points = list(pred_history)
        if pred and "scenarios" in pred:
            next_date = datetime.now() + timedelta(days=1)
            while next_date.weekday() >= 5:
                next_date += timedelta(days=1)
            all_pred_points.append({
                "date":  next_date,
                "best":  pred["scenarios"]["best_case"]["close"],
                "avg":   pred["scenarios"]["average_case"]["close"],
                "worst": pred["scenarios"]["worst_case"]["close"],
            })

        pred_dates  = []
        pred_best   = []
        pred_avg    = []
        pred_worst  = []

        if all_pred_points:
            for p in all_pred_points:
                pred_dates.append(p["date"])
                pred_best.append(p.get("best",  p.get("close", 0)))
                pred_avg.append( p.get("avg",   p.get("close", 0)))
                pred_worst.append(p.get("worst", p.get("close", 0)))

            ax_pred.fill_between(pred_dates, pred_worst, pred_best,
                                 color="#ff6b35", alpha=0.15, label="Best–Worst Range")
            ax_pred.plot(pred_dates, pred_best,  color="#2e7d32", linewidth=1.2,
                         linestyle="--", label="Best Case")
            ax_pred.plot(pred_dates, pred_avg,   color="#ff6b35", linewidth=1.5,
                         label="Avg Prediction")
            ax_pred.plot(pred_dates, pred_worst, color="#c62828", linewidth=1.2,
                         linestyle="--", label="Worst Case")
            ax_pred.scatter(pred_dates, pred_avg, color="#ff6b35", s=30, zorder=5)

        else:
            ax_pred.text(0.5, 0.5, "No prediction history yet",
                         ha="center", va="center", color="gray",
                         transform=ax_pred.transAxes, fontsize=10)

        ax_pred.set_title(f"{symbol} — Prediction Scenarios", fontsize=10, pad=8)
        ax_pred.set_ylabel("Price (USD)", fontsize=9)
        ax_pred.set_xlabel("Date", fontsize=9)
        ax_pred.xaxis.set_major_formatter(mdates.DateFormatter("%b %d"))
        ax_pred.xaxis.set_major_locator(mdates.AutoDateLocator())
        plt.setp(ax_pred.xaxis.get_majorticklabels(), rotation=30, ha="right")
        ax_pred.legend(fontsize=8)

        # Crosshair for bottom chart
        vline_pred = ax_pred.axvline(x=pred_dates[0] if pred_dates else datetime.now(),
                                     color="gray", linewidth=0.8, linestyle=":", visible=False)
        hline_pred = ax_pred.axhline(y=0, color="gray", linewidth=0.8, linestyle=":", visible=False)
        dot_pred,  = ax_pred.plot([], [], "o", color="#ff6b35", markersize=6, zorder=6)
        tooltip_pred = ax_pred.annotate(
            "", xy=(0, 0), xytext=(12, 12), textcoords="offset points",
            bbox=dict(boxstyle="round,pad=0.4", fc="white", ec="#555", alpha=0.9),
            fontsize=8, visible=False
        )

        fig.tight_layout(pad=1.5)

        canvas = FigureCanvasTkAgg(fig, master=frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        self._chart_canvas[symbol] = canvas

        # ── Hover handler ──────────────────────────────────────────── #
        import numpy as np

        def on_move(event):
            # ---- top chart ----
            if event.inaxes == ax_main and actual_dates:
                x_num = mdates.date2num(actual_dates)
                mx    = event.xdata
                if mx is None:
                    return
                idx = int(np.argmin(np.abs(x_num - mx)))
                idx = max(0, min(idx, len(actual_dates) - 1))
                xv, yv = actual_dates[idx], actual_closes[idx]
                vline_main.set_xdata([xv, xv]); vline_main.set_visible(True)
                hline_main.set_ydata([yv, yv]); hline_main.set_visible(True)
                dot_main.set_data([xv], [yv])
                label = f"{xv.strftime('%Y-%m-%d') if hasattr(xv,'strftime') else str(xv)[:10]}\n${yv:.2f}"
                tooltip_main.set_text(label)
                tooltip_main.xy = (xv, yv)
                tooltip_main.set_visible(True)
            else:
                vline_main.set_visible(False)
                hline_main.set_visible(False)
                dot_main.set_data([], [])
                tooltip_main.set_visible(False)

            # ---- bottom chart ----
            if event.inaxes == ax_pred and pred_dates:
                x_num = mdates.date2num(pred_dates)
                mx    = event.xdata
                if mx is None:
                    canvas.draw_idle()
                    return
                idx = int(np.argmin(np.abs(x_num - mx)))
                idx = max(0, min(idx, len(pred_dates) - 1))
                xv  = pred_dates[idx]
                yv  = pred_avg[idx]
                yb  = pred_best[idx]
                yw  = pred_worst[idx]
                vline_pred.set_xdata([xv, xv]); vline_pred.set_visible(True)
                hline_pred.set_ydata([yv, yv]); hline_pred.set_visible(True)
                dot_pred.set_data([xv], [yv])
                dt_str = xv.strftime("%Y-%m-%d") if hasattr(xv, "strftime") else str(xv)[:10]
                label = (f"{dt_str}\nAvg: ${yv:.2f}"
                         f"\nBest: ${yb:.2f}\nWorst: ${yw:.2f}")
                tooltip_pred.set_text(label)
                tooltip_pred.xy = (xv, yv)
                tooltip_pred.set_visible(True)
            else:
                vline_pred.set_visible(False)
                hline_pred.set_visible(False)
                dot_pred.set_data([], [])
                tooltip_pred.set_visible(False)

            canvas.draw_idle()

        canvas.mpl_connect("motion_notify_event", on_move)
        plt.close(fig)

    def refresh_all_charts(self):
        for symbol in self.stocks:
            self._draw_chart(symbol)
        self.log("Charts refreshed")

    def _remove_chart_tab(self, symbol):
        if symbol in self._chart_tabs:
            frame = self._chart_tabs.pop(symbol)
            idx = self.chart_nb.index(frame)
            self.chart_nb.forget(idx)
        self._chart_canvas.pop(symbol, None)

    # ------------------------------------------------------------------ #
    #  STOCK MANAGEMENT                                                    #
    # ------------------------------------------------------------------ #

    def add_stock(self):
        symbol = self.symbol_var.get().strip().upper()
        if not symbol:
            messagebox.showwarning("Warning", "Enter a stock symbol.")
            return
        if symbol in self.stocks:
            messagebox.showwarning("Warning", f"{symbol} is already tracked.")
            return
        self.stocks[symbol] = {
            "system": None, "prediction": None, "raw_df": None,
            "pred_history": [],
            "status": "Training…",
            "lookback": self.lookback_var.get(),
            "epochs":   self.epochs_var.get(),
        }
        self._update_tree_row(symbol)
        self._ensure_chart_tab(symbol)
        threading.Thread(target=self._train_thread, args=(symbol,), daemon=True).start()
        self.log(f"Queued training for {symbol}")
        self.symbol_var.set("")

    def quick_add(self):
        for sym in ["SPY", "AAPL", "MSFT"]:
            if sym not in self.stocks:
                self.symbol_var.set(sym)
                self.add_stock()

    def remove_selected(self):
        syms = self._selected_symbols()
        if not syms:
            messagebox.showwarning("Warning", "Select a stock first.")
            return
        if messagebox.askyesno("Confirm", f"Remove {', '.join(syms)}?"):
            for sym in syms:
                if self._tree_has(sym):
                    self.tree.delete(sym)
                self.stocks.pop(sym, None)
                self._remove_chart_tab(sym)
                self.log(f"Removed {sym}")

    # ------------------------------------------------------------------ #
    #  BACKGROUND THREADS                                                  #
    # ------------------------------------------------------------------ #

    def _train_thread(self, symbol):
        try:
            data = self.stocks[symbol]
            system = StockTradingSystem(api_key="", lookback_window=data["lookback"])
            self.message_queue.put(("status", f"Training {symbol}…"))

            end_date   = datetime.now().strftime("%Y-%m-%d")
            start_date = (datetime.now() - timedelta(days=180)).strftime("%Y-%m-%d")
            raw_df = system.api.fetch_stock_data(symbol, start_date, end_date)
            data["raw_df"] = raw_df

            system.train_model(symbol, epochs=data["epochs"])
            data["system"] = system
            data["status"] = "Trained"

            pred = system.predict_next_day(symbol, include_scenarios=True)
            data["prediction"] = pred
            data["status"] = "Ready"

            self.message_queue.put(("refresh", symbol))
            self.message_queue.put(("chart",   symbol))
            self.message_queue.put(("log",     f"✓ {symbol} trained and predicted"))
            self.message_queue.put(("status",  f"Ready — {symbol} complete"))
        except Exception as e:
            err = str(e)
            if symbol in self.stocks:
                self.stocks[symbol]["status"] = f"Error: {err[:40]}"
            self.message_queue.put(("refresh", symbol))
            self.message_queue.put(("log",     f"✗ {symbol}: {err}"))

    def _predict_thread(self, symbol):
        try:
            data = self.stocks[symbol]
            if not data["system"]:
                self.message_queue.put(("log", f"{symbol} not yet trained."))
                return
            self.message_queue.put(("status", f"Predicting {symbol}…"))
            pred = data["system"].predict_next_day(symbol, include_scenarios=True)

            # Archive previous prediction into history before overwriting
            old_pred = data.get("prediction")
            if old_pred and "scenarios" in old_pred:
                data["pred_history"].append({
                    "date":  datetime.now(),
                    "avg":   old_pred["scenarios"]["average_case"]["close"],
                    "best":  old_pred["scenarios"]["best_case"]["close"],
                    "worst": old_pred["scenarios"]["worst_case"]["close"],
                })

            data["prediction"] = pred
            data["status"] = "Ready"
            self.message_queue.put(("refresh", symbol))
            self.message_queue.put(("chart",   symbol))
            self.message_queue.put(("log",     f"✓ {symbol} predictions refreshed"))
            self.message_queue.put(("status",  "Ready"))
        except Exception as e:
            self.message_queue.put(("log", f"✗ {symbol} predict error: {e}"))

    def _update_thread(self, symbol):
        try:
            data = self.stocks[symbol]
            if not data["system"]:
                self.message_queue.put(("log", f"{symbol} not yet trained."))
                return
            self.message_queue.put(("status", f"Updating {symbol}…"))
            data["system"].adaptive_update(symbol)

            old_pred = data.get("prediction")
            if old_pred and "scenarios" in old_pred:
                data["pred_history"].append({
                    "date":  datetime.now(),
                    "avg":   old_pred["scenarios"]["average_case"]["close"],
                    "best":  old_pred["scenarios"]["best_case"]["close"],
                    "worst": old_pred["scenarios"]["worst_case"]["close"],
                })

            pred = data["system"].predict_next_day(symbol, include_scenarios=True)
            data["prediction"] = pred
            data["status"] = "Ready"
            self.message_queue.put(("refresh", symbol))
            self.message_queue.put(("chart",   symbol))
            self.message_queue.put(("log",     f"✓ {symbol} adaptively updated"))
            self.message_queue.put(("status",  "Ready"))
        except Exception as e:
            self.message_queue.put(("log", f"✗ {symbol} update error: {e}"))

    # ------------------------------------------------------------------ #
    #  BUTTON ACTIONS                                                      #
    # ------------------------------------------------------------------ #

    def predict_selected(self):
        syms = self._selected_symbols()
        if not syms:
            messagebox.showwarning("Warning", "Select a stock first.")
            return
        for sym in syms:
            threading.Thread(target=self._predict_thread, args=(sym,), daemon=True).start()

    def predict_all(self):
        for sym in list(self.stocks):
            threading.Thread(target=self._predict_thread, args=(sym,), daemon=True).start()

    def update_selected(self):
        syms = self._selected_symbols()
        if not syms:
            messagebox.showwarning("Warning", "Select a stock first.")
            return
        for sym in syms:
            threading.Thread(target=self._update_thread, args=(sym,), daemon=True).start()

    def update_all(self):
        for sym in list(self.stocks):
            threading.Thread(target=self._update_thread, args=(sym,), daemon=True).start()

    # ------------------------------------------------------------------ #
    #  EXCEL HELPERS                                                       #
    # ------------------------------------------------------------------ #

    def _df_for_export(self, symbol):
        """Return a clean OHLCV DataFrame ready to write."""
        df = self.stocks[symbol].get("raw_df")
        if df is None:
            return None
        df_out = df[["open", "high", "low", "close", "volume"]].copy()
        df_out.columns = ["Open", "High", "Low", "Close", "Volume"]
        if hasattr(df_out.index, "tz") and df_out.index.tz is not None:
            df_out.index = df_out.index.tz_localize(None)
        df_out.index = df_out.index.date
        df_out.index.name = "Date"
        return df_out

    def _pred_df_for_export(self, symbol):
        """Return a predictions DataFrame for one symbol."""
        pred = self.stocks[symbol].get("prediction")
        if not pred or "scenarios" not in pred:
            return None
        sc = pred["scenarios"]
        rows = []
        for label, key in [("Best Case",    "best_case"),
                            ("Average Case", "average_case"),
                            ("Worst Case",   "worst_case")]:
            s = sc[key]
            rows.append({
                "Exported At":   datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Scenario":      label,
                "Open":          round(s["open"],  2),
                "High":          round(s["high"],  2),
                "Low":           round(s["low"],   2),
                "Close":         round(s["close"], 2),
                "Profit %":      round(s["profit_potential"], 2),
                "Current Price": round(pred["current_price"], 2),
                "Confidence":    round(pred["confidence"] * 100, 1),
                "Signal":        pred.get("recommendation", "").replace("_", " "),
            })
        return pd.DataFrame(rows)

    # ------------------------------------------------------------------ #
    #  EXPORT — STOCK DATA (fresh file)                                    #
    # ------------------------------------------------------------------ #

    def export_stock_data(self):
        has_data = any(d.get("raw_df") is not None for d in self.stocks.values())
        if not has_data:
            messagebox.showinfo("Info", "No stock data yet. Train a stock first.")
            return
        try:
            with pd.ExcelWriter(STOCK_DATA_FILE, engine="openpyxl") as writer:
                for symbol in self.stocks:
                    df_out = self._df_for_export(symbol)
                    if df_out is not None:
                        df_out.to_excel(writer, sheet_name=symbol)
            self.log(f"Stock data exported → {STOCK_DATA_FILE}")
            messagebox.showinfo("Exported", f"Saved as:\n{os.path.abspath(STOCK_DATA_FILE)}")
        except PermissionError:
            messagebox.showerror("Error", f"{STOCK_DATA_FILE} is open. Close it and retry.")

    # ------------------------------------------------------------------ #
    #  UPDATE — STOCK DATA (append new rows, keep history)                #
    # ------------------------------------------------------------------ #

    def update_stock_data(self):
        has_data = any(d.get("raw_df") is not None for d in self.stocks.values())
        if not has_data:
            messagebox.showinfo("Info", "No stock data yet. Train a stock first.")
            return
        try:
            if not os.path.exists(STOCK_DATA_FILE):
                self.export_stock_data()
                return

            from openpyxl import load_workbook
            wb = load_workbook(STOCK_DATA_FILE)

            for symbol in self.stocks:
                df_new = self._df_for_export(symbol)
                if df_new is None:
                    continue

                if symbol in wb.sheetnames:
                    # Read existing sheet to find last date
                    existing = pd.read_excel(
                        STOCK_DATA_FILE, sheet_name=symbol,
                        index_col=0, engine="openpyxl"
                    )
                    existing.index = pd.to_datetime(existing.index).date

                    # Keep only rows newer than what we already have
                    last_date = existing.index.max()
                    df_append = df_new[df_new.index > last_date]

                    if df_append.empty:
                        self.log(f"{symbol}: no new rows to append")
                        continue

                    df_combined = pd.concat([existing, df_append])
                else:
                    df_combined = df_new

                # Rewrite that sheet
                if symbol in wb.sheetnames:
                    del wb[symbol]
                wb.save(STOCK_DATA_FILE)

                # Append the combined df using ExcelWriter in append mode
                with pd.ExcelWriter(
                    STOCK_DATA_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace"
                ) as writer:
                    df_combined.to_excel(writer, sheet_name=symbol)

                self.log(f"{symbol}: stock data updated in {STOCK_DATA_FILE}")

            messagebox.showinfo("Updated", f"Stock data updated in:\n{os.path.abspath(STOCK_DATA_FILE)}")
        except PermissionError:
            messagebox.showerror("Error", f"{STOCK_DATA_FILE} is open. Close it and retry.")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.log(f"Update stock data error: {e}")

    # ------------------------------------------------------------------ #
    #  EXPORT — PREDICTIONS (fresh file)                                  #
    # ------------------------------------------------------------------ #

    def export_predictions(self):
        has_preds = any(d.get("prediction") is not None for d in self.stocks.values())
        if not has_preds:
            messagebox.showinfo("Info", "No predictions yet. Train and predict first.")
            return
        try:
            with pd.ExcelWriter(PREDICTIONS_FILE, engine="openpyxl") as writer:
                for symbol in self.stocks:
                    df_pred = self._pred_df_for_export(symbol)
                    if df_pred is not None:
                        df_pred.to_excel(writer, sheet_name=symbol, index=False)
            self.log(f"Predictions exported → {PREDICTIONS_FILE}")
            messagebox.showinfo("Exported", f"Saved as:\n{os.path.abspath(PREDICTIONS_FILE)}")
        except PermissionError:
            messagebox.showerror("Error", f"{PREDICTIONS_FILE} is open. Close it and retry.")

    # ------------------------------------------------------------------ #
    #  UPDATE — PREDICTIONS (append new prediction rows, keep history)    #
    # ------------------------------------------------------------------ #

    def update_predictions(self):
        has_preds = any(d.get("prediction") is not None for d in self.stocks.values())
        if not has_preds:
            messagebox.showinfo("Info", "No predictions yet. Train and predict first.")
            return
        try:
            if not os.path.exists(PREDICTIONS_FILE):
                self.export_predictions()
                return

            for symbol in self.stocks:
                df_new = self._pred_df_for_export(symbol)
                if df_new is None:
                    continue

                if os.path.exists(PREDICTIONS_FILE):
                    try:
                        existing = pd.read_excel(
                            PREDICTIONS_FILE, sheet_name=symbol,
                            engine="openpyxl"
                        )
                        df_combined = pd.concat([existing, df_new], ignore_index=True)
                    except Exception:
                        # Sheet doesn't exist yet
                        df_combined = df_new
                else:
                    df_combined = df_new

                with pd.ExcelWriter(
                    PREDICTIONS_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace"
                ) as writer:
                    df_combined.to_excel(writer, sheet_name=symbol, index=False)

                self.log(f"{symbol}: predictions appended to {PREDICTIONS_FILE}")

            messagebox.showinfo("Updated", f"Predictions updated in:\n{os.path.abspath(PREDICTIONS_FILE)}")
        except PermissionError:
            messagebox.showerror("Error", f"{PREDICTIONS_FILE} is open. Close it and retry.")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.log(f"Update predictions error: {e}")

    # ------------------------------------------------------------------ #
    #  MESSAGE QUEUE                                                       #
    # ------------------------------------------------------------------ #

    def process_queue(self):
        try:
            while True:
                msg_type, payload = self.message_queue.get_nowait()
                if msg_type == "log":
                    self.log(payload)
                elif msg_type == "status":
                    self.status_var.set(payload)
                elif msg_type == "refresh":
                    self._update_tree_row(payload)
                elif msg_type == "chart":
                    self._draw_chart(payload)
        except queue.Empty:
            pass
        self.root.after(100, self.process_queue)


def main():
    root = tk.Tk()
    StockPriceGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
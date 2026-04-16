# Excel I/O for OHLCV data, predictions, and daily scores — no ML, no registry logic.

from __future__ import annotations

import os
import threading
from datetime import datetime
from typing import Dict, List, Optional

import pandas as pd

from scoring.scorer import _parse_records, _match_actuals, _prev_close

_LOCK = threading.Lock()

STOCK_DATA_FILE  = "stock_data.xlsx"
PREDICTIONS_FILE = "stock_predictions.xlsx"
SCORES_FILE      = "prediction_score.xlsx"


# Exports OHLCV history, prediction scenarios, and daily scores to Excel files.
class ExcelExporter:

    def __init__(
        self,
        stock_data_file: str = STOCK_DATA_FILE,
        predictions_file: str = PREDICTIONS_FILE,
        scores_file: str = SCORES_FILE,
    ) -> None:
        self._data_file  = stock_data_file
        self._pred_file  = predictions_file
        self._score_file = scores_file

    # ------------------------------------------------------------------ #
    #  OHLCV data                                                         #
    # ------------------------------------------------------------------ #

    # Write all stocks' OHLCV data to Excel (one sheet per symbol), overwriting the file.
    def export_stock_data(self, stocks: Dict) -> str:
        with _LOCK:
            with pd.ExcelWriter(self._data_file, engine="openpyxl") as writer:
                for symbol, data in stocks.items():
                    df_out = self._df_for_export(data)
                    if df_out is not None:
                        df_out.to_excel(writer, sheet_name=symbol)
        return os.path.abspath(self._data_file)

    # Append only new rows since the last export; creates the file if it doesn't exist.
    def update_stock_data(self, stocks: Dict) -> str:
        if not os.path.exists(self._data_file):
            return self.export_stock_data(stocks)

        from openpyxl import load_workbook
        try:
            wb = load_workbook(self._data_file)
        except Exception:
            return self.export_stock_data(stocks)

        with _LOCK:
            for symbol, data in stocks.items():
                df_new = self._df_for_export(data)
                if df_new is None:
                    continue
                if symbol in wb.sheetnames:
                    existing = pd.read_excel(
                        self._data_file, sheet_name=symbol,
                        index_col=0, engine="openpyxl",
                    )
                    existing.index = pd.to_datetime(existing.index).date
                    last_date  = existing.index.max()
                    df_append  = df_new[df_new.index > last_date]
                    if df_append.empty:
                        continue
                    df_combined = pd.concat([existing, df_append])
                    del wb[symbol]
                    wb.save(self._data_file)
                else:
                    df_combined = df_new
                with pd.ExcelWriter(
                    self._data_file, engine="openpyxl",
                    mode="a", if_sheet_exists="replace",
                ) as writer:
                    df_combined.to_excel(writer, sheet_name=symbol)

        return os.path.abspath(self._data_file)

    # ------------------------------------------------------------------ #
    #  Predictions                                                        #
    # ------------------------------------------------------------------ #

    # Write all stocks' prediction scenarios to Excel (one sheet per symbol), overwriting the file.
    def export_predictions(self, stocks: Dict) -> str:
        with _LOCK:
            with pd.ExcelWriter(self._pred_file, engine="openpyxl") as writer:
                for symbol, data in stocks.items():
                    df_pred = self._pred_df_for_export(symbol, data)
                    if df_pred is not None:
                        df_pred.to_excel(writer, sheet_name=symbol, index=False)
        return os.path.abspath(self._pred_file)

    # Append new prediction rows to each symbol's sheet; creates the file if it doesn't exist.
    def update_predictions(self, stocks: Dict) -> str:
        if not os.path.exists(self._pred_file):
            return self.export_predictions(stocks)

        with _LOCK:
            for symbol, data in stocks.items():
                df_new = self._pred_df_for_export(symbol, data)
                if df_new is None:
                    continue
                try:
                    existing    = pd.read_excel(self._pred_file, sheet_name=symbol, engine="openpyxl")
                    df_combined = pd.concat([existing, df_new], ignore_index=True)
                except Exception:
                    df_combined = df_new
                with pd.ExcelWriter(
                    self._pred_file, engine="openpyxl",
                    mode="a", if_sheet_exists="replace",
                ) as writer:
                    df_combined.to_excel(writer, sheet_name=symbol, index=False)

        return os.path.abspath(self._pred_file)

    # ------------------------------------------------------------------ #
    #  Private helpers                                                    #
    # ------------------------------------------------------------------ #

    # Extract and clean the OHLCV DataFrame from a stock entry for Excel export (excludes today).
    @staticmethod
    def _df_for_export(data: Dict) -> Optional[pd.DataFrame]:
        df = data.get("raw_df")
        if df is None:
            return None
        df_out = df[["open", "high", "low", "close", "volume"]].copy()
        df_out.columns = ["Open", "High", "Low", "Close", "Volume"]
        if hasattr(df_out.index, "tz") and df_out.index.tz is not None:
            df_out.index = df_out.index.tz_localize(None)
        df_out.index = df_out.index.date
        df_out.index.name = "Date"
        today = datetime.now().date()
        df_out = df_out[df_out.index < today]
        return df_out

    # Reconstruct pred_history from Excel by grouping Best/Average/Worst rows by their export timestamp.
    def load_pred_history(self, symbol: str) -> List[Dict]:
        if not os.path.exists(self._pred_file):
            return []
        try:
            df = pd.read_excel(self._pred_file, sheet_name=symbol, engine="openpyxl")
        except Exception:
            return []

        scenario_rows = df[df["Scenario"].isin(["Best Case", "Average Case", "Worst Case"])]
        if scenario_rows.empty:
            return []

        groups: Dict[str, Dict] = {}
        for _, row in scenario_rows.iterrows():
            key      = str(row["Exported At"])
            scenario = str(row["Scenario"])
            try:
                close = float(row["Close"])
            except (ValueError, TypeError):
                continue
            groups.setdefault(key, {})[scenario] = close

        records = []
        for exported_at, scenarios in groups.items():
            if "Average Case" not in scenarios:
                continue
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
                try:
                    dt = datetime.strptime(exported_at[:19], fmt) if len(fmt) == 19 else \
                         datetime.strptime(exported_at[:16], fmt) if len(fmt) == 16 else \
                         datetime.strptime(exported_at[:10], fmt)
                    break
                except ValueError:
                    continue
            else:
                continue
            records.append({
                "date":  dt,
                "avg":   scenarios["Average Case"],
                "best":  scenarios.get("Best Case",  scenarios["Average Case"]),
                "worst": scenarios.get("Worst Case", scenarios["Average Case"]),
            })

        return records

    # Build the predictions DataFrame (scenario rows only) ready for Excel export.
    @staticmethod
    def _pred_df_for_export(symbol: str, data: Dict) -> Optional[pd.DataFrame]:
        pred = data.get("prediction")
        if not pred or "scenarios" not in pred:
            return None

        sc   = pred["scenarios"]
        rows = []
        for label, key in [
            ("Best Case",    "best_case"),
            ("Average Case", "average_case"),
            ("Worst Case",   "worst_case"),
        ]:
            s = sc[key]
            rows.append({
                "Exported At":    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Scenario":       label,
                "Open":           round(s["open"],  2),
                "High":           round(s["high"],  2),
                "Low":            round(s["low"],   2),
                "Close":          round(s["close"], 2),
                "Profit %":       round(s["profit_potential"], 2),
                "Current Price":  round(pred["current_price"], 2),
                "Confidence":     round(pred["confidence"] * 100, 1),
                "Signal":         pred.get("recommendation", "").replace("_", " "),
            })

        return pd.DataFrame(rows)

    # ------------------------------------------------------------------ #
    #  Scores (prediction_score.xlsx)                                     #
    # ------------------------------------------------------------------ #

    # Write per-symbol daily score sheets to prediction_score.xlsx, overwriting the file.
    def export_scores(self, stocks: Dict) -> str:
        with _LOCK:
            with pd.ExcelWriter(self._score_file, engine="openpyxl") as writer:
                for symbol, data in stocks.items():
                    df_score = self._score_df_for_export(symbol, data)
                    if df_score is not None:
                        df_score.to_excel(writer, sheet_name=symbol, index=False)
        return os.path.abspath(self._score_file)

    # Create prediction_score.xlsx if it doesn't exist yet, then overwrite with the latest score data.
    def update_scores(self, stocks: Dict) -> str:
        if not os.path.exists(self._score_file):
            self._create_empty_score_file()
        return self.export_scores(stocks)

    # Read "── Daily Score ──" rows out of stock_predictions.xlsx, write them to prediction_score.xlsx,
    # and strip them from stock_predictions.xlsx.  Safe to call multiple times (idempotent).
    # All reads happen before any writes so neither file is ever left in a partial state.
    def migrate_scores_from_predictions(self) -> int:
        """Returns the total number of score rows migrated across all symbols."""
        if not os.path.exists(self._pred_file):
            return 0

        # ── Phase 1: read everything into memory ──────────────────────── #
        try:
            pred_xl = pd.ExcelFile(self._pred_file, engine="openpyxl")
        except Exception:
            return 0

        pred_sheets: Dict[str, pd.DataFrame] = {}
        for sheet_name in pred_xl.sheet_names:
            try:
                pred_sheets[str(sheet_name)] = pred_xl.parse(sheet_name)
            except Exception:
                pass
        pred_xl.close()

        score_sheets: Dict[str, pd.DataFrame] = {}
        if os.path.exists(self._score_file):
            try:
                score_xl = pd.ExcelFile(self._score_file, engine="openpyxl")
                for sheet_name in score_xl.sheet_names:
                    try:
                        df = score_xl.parse(sheet_name)
                        if not df.empty and "Prediction Date" in df.columns:
                            score_sheets[str(sheet_name)] = df
                    except Exception:
                        pass
                score_xl.close()
            except Exception:
                pass

        # ── Phase 2: process score rows in memory ─────────────────────── #
        total_migrated  = 0
        cleaned_pred: Dict[str, pd.DataFrame] = {}
        pred_file_dirty = False

        for sheet_name, df in pred_sheets.items():
            if "Scenario" not in df.columns:
                cleaned_pred[sheet_name] = df
                continue

            score_mask = df["Scenario"] == "\u2500\u2500 Daily Score \u2500\u2500"
            score_rows = df[score_mask]

            if score_rows.empty:
                cleaned_pred[sheet_name] = df
                continue

            new_rows: List[Dict] = []
            for _, row in score_rows.iterrows():
                pred_date = str(row.get("Exported At", ""))
                close_str = str(row.get("Close", ""))
                actual    = float(row.get("Current Price", 0) or 0)
                signal    = str(row.get("Signal", ""))
                in_range  = "\u2713" if "\u2713" in signal else "\u2717"

                predicted = actual
                if "\u2192" in close_str:
                    try:
                        predicted = float(close_str.split("\u2192")[0].replace("Pred", "").strip())
                    except ValueError:
                        pass

                err_pct = abs(predicted - actual) / (abs(actual) + 1e-8) * 100
                new_rows.append({
                    "Prediction Date": pred_date,
                    "Predicted Close": round(predicted, 2),
                    "Best Case":       round(predicted, 2),
                    "Worst Case":      round(predicted, 2),
                    "Actual Close":    round(actual,    2),
                    "Error %":         round(err_pct,   2),
                    "In Range":        in_range,
                    "Direction":       "N/A",
                })

            df_new = pd.DataFrame(new_rows)

            if sheet_name in score_sheets:
                df_merged = pd.concat([score_sheets[sheet_name], df_new], ignore_index=True)
                df_merged = df_merged.drop_duplicates(subset=["Prediction Date"], keep="last")
            else:
                df_merged = df_new

            score_sheets[sheet_name] = df_merged
            total_migrated += len(new_rows)
            cleaned_pred[sheet_name] = df[~score_mask].reset_index(drop=True)
            pred_file_dirty = True

        # ── Phase 3: write both files (no partial saves) ──────────────── #
        with _LOCK:
            if pred_file_dirty:
                with pd.ExcelWriter(self._pred_file, engine="openpyxl") as writer:
                    for sym, df_sym in cleaned_pred.items():
                        df_sym.to_excel(writer, sheet_name=sym, index=False)

            if score_sheets:
                with pd.ExcelWriter(self._score_file, engine="openpyxl") as writer:
                    for sym, df_sym in score_sheets.items():
                        df_sym.to_excel(writer, sheet_name=sym, index=False)
            elif not os.path.exists(self._score_file):
                self._create_empty_score_file()

        return total_migrated

    # Create an empty prediction_score.xlsx as a placeholder before any predictions are matched.
    def _create_empty_score_file(self) -> None:
        from openpyxl import Workbook
        with _LOCK:
            wb = Workbook()
            wb.active.title = "Score Data"  # type: ignore[union-attr]
            wb.save(self._score_file)

    # Build a per-symbol DataFrame of matched predictions vs actuals for the scores file.
    @staticmethod
    def _score_df_for_export(symbol: str, data: Dict) -> Optional[pd.DataFrame]:
        ph = data.get("pred_history", [])
        df = data.get("raw_df")
        if not ph or df is None:
            return None

        records = _parse_records(ph)
        records = _match_actuals(records, df)
        matched = [r for r in records if r.actual is not None]
        if not matched:
            return None

        rows: List[Dict] = []
        for r in matched:
            avg    = float(r.avg)
            best   = float(r.best)
            worst  = float(r.worst)
            actual = float(r.actual)  # type: ignore[arg-type]
            err_pct  = abs(avg - actual) / (abs(actual) + 1e-8) * 100
            in_range = min(best, worst) <= actual <= max(best, worst)
            prev      = _prev_close(r.date, df)
            direction = "N/A"
            if prev is not None:
                direction = "✓" if (avg >= prev) == (actual >= prev) else "✗"
            pred_date = r.date.strftime("%Y-%m-%d %H:%M") if hasattr(r.date, "strftime") else str(r.date)
            rows.append({
                "Prediction Date": pred_date,
                "Predicted Close": round(avg,    2),
                "Best Case":       round(best,   2),
                "Worst Case":      round(worst,  2),
                "Actual Close":    round(actual, 2),
                "Error %":         round(err_pct, 2),
                "In Range":        "✓" if in_range else "✗",
                "Direction":       direction,
            })

        return pd.DataFrame(rows)

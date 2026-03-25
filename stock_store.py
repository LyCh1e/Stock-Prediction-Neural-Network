"""
stock_store.py
~~~~~~~~~~~~~~
Data-management layer for the Stock Price Predictor.

Responsibilities
----------------
- In-memory stock registry (add / remove / get)
- Symbol persistence  → tracked_symbols.json
- Excel export/update → stock_data.xlsx, stock_predictions.xlsx, stock_models.xlsx
- DataFrame helpers   → OHLCV and predictions frames
- Background thread workers (train / predict / adaptive-update)

The GUI imports StockStore and delegates all data work here, keeping
stock_gui.py focused purely on presentation.
"""

from __future__ import annotations

import json
import math
import os
import threading
from datetime import datetime, timedelta
from typing import Callable, Dict, Optional

import numpy as np

import pandas as pd

from stock_volume_predictor import StockTradingSystem
from stock_scorer import score_symbol, score_all, ScoreResult

# ── File name constants ──────────────────────────────────────────────────── #
STOCK_DATA_FILE  = "stock_data.xlsx"
PREDICTIONS_FILE = "stock_predictions.xlsx"
MODELS_FILE      = "stock_models.xlsx"
SYMBOLS_FILE     = "tracked_symbols.json"

# ── Type alias ───────────────────────────────────────────────────────────── #
StockEntry = Dict


class StockStore:
    """
    Central data store for tracked stocks.

    Parameters
    ----------
    message_cb : callable
        Callback used to post messages back to the GUI.
        Signature: ``message_cb(msg_type: str, payload)``
        msg_type values: "log" | "status" | "refresh" | "chart"
    """

    def __init__(self, message_cb: Callable[[str, object], None]) -> None:
        self._stocks: Dict[str, StockEntry] = {}
        self._cb = message_cb

    # ------------------------------------------------------------------ #
    #  Public registry accessors                                           #
    # ------------------------------------------------------------------ #

    @property
    def stocks(self) -> Dict[str, StockEntry]:
        """Read-only view of the in-memory stock registry."""
        return self._stocks

    def get(self, symbol: str) -> Optional[StockEntry]:
        return self._stocks.get(symbol)

    def symbols(self):
        return list(self._stocks.keys())

    def has(self, symbol: str) -> bool:
        return symbol in self._stocks

    # ------------------------------------------------------------------ #
    #  Create / Delete                                                     #
    # ------------------------------------------------------------------ #

    def add(self, symbol: str, lookback: int = 10, epochs: int = 200) -> bool:
        """
        Register a new stock and kick off background training.

        Returns True if the stock was added, False if it already existed.
        """
        symbol = symbol.upper()
        if symbol in self._stocks:
            return False

        self._stocks[symbol] = {
            "system":         None,
            "prediction":     None,
            "raw_df":         None,
            "pred_history":   [],
            "accuracy_score": None,
            "status":         "Training…",
            "lookback":       lookback,
            "epochs":         epochs,
        }
        self.save_symbols()
        threading.Thread(
            target=self._train_thread, args=(symbol,), daemon=True
        ).start()
        return True

    def remove(self, symbol: str) -> bool:
        """
        Remove a stock from the registry.

        Returns True if removed, False if it wasn't tracked.
        """
        symbol = symbol.upper()
        if symbol not in self._stocks:
            return False
        del self._stocks[symbol]
        self.save_symbols()
        return True

    # ------------------------------------------------------------------ #
    #  Trigger background operations                                       #
    # ------------------------------------------------------------------ #

    def predict(self, symbol: str) -> None:
        """Kick off a prediction refresh for *symbol* in a background thread."""
        threading.Thread(
            target=self._predict_thread, args=(symbol,), daemon=True
        ).start()

    def update(self, symbol: str) -> None:
        """Kick off an adaptive model update for *symbol* in a background thread."""
        threading.Thread(
            target=self._update_thread, args=(symbol,), daemon=True
        ).start()

    # ------------------------------------------------------------------ #
    #  Symbol persistence                                                  #
    # ------------------------------------------------------------------ #

    def save_symbols(self) -> None:
        """Persist tracked symbols and their settings to *SYMBOLS_FILE*."""
        try:
            data = {
                sym: {"lookback": info["lookback"], "epochs": info["epochs"]}
                for sym, info in self._stocks.items()
            }
            with open(SYMBOLS_FILE, "w") as fh:
                json.dump(data, fh, indent=2)
            self._cb("log", f"Symbols saved → {SYMBOLS_FILE}")
        except Exception as exc:
            self._cb("log", f"Warning: could not save symbols file: {exc}")

    def load_symbols(self) -> None:
        """
        Restore tracked symbols from *SYMBOLS_FILE* on startup.
        Each restored symbol triggers a fresh training run.
        """
        if not os.path.exists(SYMBOLS_FILE):
            return
        try:
            with open(SYMBOLS_FILE, "r") as fh:
                data: dict = json.load(fh)
            if not data:
                return
            self._cb("log", f"Loading {len(data)} saved symbol(s) from {SYMBOLS_FILE}…")
            for symbol, settings in data.items():
                symbol = symbol.upper()
                if symbol in self._stocks:
                    continue
                self._stocks[symbol] = {
                    "system":         None,
                    "prediction":     None,
                    "raw_df":         None,
                    "pred_history":   [],
                    "accuracy_score": None,
                    "status":         "Training…",
                    "lookback":       settings.get("lookback", 10),
                    "epochs":         settings.get("epochs", 200),
                }
                self._cb("refresh", symbol)
                self._cb("chart",   symbol)
                threading.Thread(
                    target=self._train_thread, args=(symbol,), daemon=True
                ).start()
                self._cb("log", f"Queued training for {symbol} (restored from save)")
        except Exception as exc:
            self._cb("log", f"Warning: could not load symbols file: {exc}")

    # ------------------------------------------------------------------ #
    #  DataFrame helpers                                                   #
    # ------------------------------------------------------------------ #

    def df_for_export(self, symbol: str) -> Optional[pd.DataFrame]:
        """Return a clean OHLCV DataFrame ready to write to Excel."""
        df = self._stocks[symbol].get("raw_df")
        if df is None:
            return None
        df_out = df[["open", "high", "low", "close", "volume"]].copy()
        df_out.columns = ["Open", "High", "Low", "Close", "Volume"]
        if hasattr(df_out.index, "tz") and df_out.index.tz is not None:
            df_out.index = df_out.index.tz_localize(None)
        df_out.index = df_out.index.date
        df_out.index.name = "Date"
        return df_out

    def pred_df_for_export(self, symbol: str) -> Optional[pd.DataFrame]:
        """
        Return a predictions DataFrame for *symbol*.

        Rows
        ----
        1–3  : Best / Average / Worst scenario for the *next* day prediction.
        4+   : One "Daily Score" row per archived prediction that has been
               matched to a real closing price, showing cumulative accuracy
               metrics that update day-by-day.
        """
        pred = self._stocks[symbol].get("prediction")
        if not pred or "scenarios" not in pred:
            return None

        sc = pred["scenarios"]
        rows = []
        for label, key in [
            ("Best Case",    "best_case"),
            ("Average Case", "average_case"),
            ("Worst Case",   "worst_case"),
        ]:
            s = sc[key]
            rows.append({
                "Exported At":    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Scenario":       label,
                "Open":           round(s["open"],  2),
                "High":           round(s["high"],  2),
                "Low":            round(s["low"],   2),
                "Close":          round(s["close"], 2),
                "Profit %":       round(s["profit_potential"], 2),
                "Current Price":  round(pred["current_price"], 2),
                "Confidence":     round(pred["confidence"] * 100, 1),
                "Signal":         pred.get("recommendation", "").replace("_", " "),
                # Score columns — blank for scenario rows
                "Score":          "",
                "Grade":          "",
                "Avg Error %":    "",
                "Dir Accuracy %": "",
                "In Range %":     "",
                "Matched Preds":  "",
            })

        # ── Append daily score rows below the three scenario rows ──── #
        score_rows = self._build_daily_score_rows(symbol)
        df_pred = pd.DataFrame(rows)
        if score_rows:
            df_scores = pd.DataFrame(score_rows)
            df_pred = pd.concat([df_pred, df_scores], ignore_index=True)

        return df_pred

    # ------------------------------------------------------------------ #
    #  Daily score builder                                                 #
    # ------------------------------------------------------------------ #

    def _build_daily_score_rows(self, symbol: str) -> list:
        """
        Build one score row per archived prediction that has been matched to
        an actual closing price.

        Each row shows:
          - Prediction date
          - Predicted close vs actual close
          - Whether actual landed inside the predicted band
          - Rolling composite score (0–100) + letter grade
          - Rolling avg error %, direction accuracy %, in-range %
          - Count of matched predictions so far

        The score is *cumulative/rolling* — each row shows the running score
        up to that point so you can track model improvement over time.
        """
        data = self._stocks.get(symbol)
        if data is None:
            return []

        ph = data.get("pred_history", [])
        df = data.get("raw_df")
        if not ph or df is None:
            return []

        from stock_scorer import _parse_records, _match_actuals, _prev_close

        records = _parse_records(ph)
        records = _match_actuals(records, df)
        matched = [r for r in records if r.actual is not None]
        if not matched:
            return []

        def _grade(s: float) -> str:
            if s >= 93: return "A+"
            if s >= 87: return "A"
            if s >= 80: return "A-"
            if s >= 74: return "B+"
            if s >= 67: return "B"
            if s >= 60: return "B-"
            if s >= 54: return "C+"
            if s >= 47: return "C"
            if s >= 40: return "C-"
            if s >= 34: return "D+"
            if s >= 27: return "D"
            if s >= 20: return "D-"
            return "F"

        score_rows = []
        run_errors = []
        run_dir    = []
        run_range  = []

        for r in matched:
            if r.avg is None or r.actual is None:
                continue
            err_pct  = abs(r.avg - r.actual) / (abs(r.actual) + 1e-8) * 100
            in_range = (r.best is not None and r.worst is not None and r.actual is not None and 
                        min(r.best, r.worst) <= r.actual <= max(r.best, r.worst))
            prev     = _prev_close(r.date, df)
            dir_ok   = None
            if prev is not None:
                dir_ok = (r.avg >= prev) == (r.actual >= prev)

            run_errors.append(err_pct)
            if dir_ok is not None:
                run_dir.append(dir_ok)
            run_range.append(in_range)

            # Rolling composite score (same weights as stock_scorer)
            mean_ape   = sum(run_errors) / len(run_errors)
            mape_raw   = math.exp(-0.15 * mean_ape)
            dir_acc    = (sum(run_dir) / len(run_dir)) if run_dir else 0.5
            range_frac = sum(run_range) / len(run_range)
            raw_score  = 0.50 * mape_raw + 0.30 * dir_acc + 0.20 * range_frac
            composite  = round(min(100.0, max(0.0, raw_score * 100)), 1)

            pred_date = (r.date.strftime("%Y-%m-%d %H:%M")
                         if hasattr(r.date, "strftime") else str(r.date))

            score_rows.append({
                "Exported At":    pred_date,
                "Scenario":       "── Daily Score ──",
                "Open":           "",
                "High":           "",
                "Low":            "",
                "Close":          f"Pred {round(r.avg or 0, 2)}  →  Actual {round(r.actual or 0, 2)}",
                "Profit %":       round((r.actual - r.avg) / (abs(r.avg) + 1e-8) * 100, 2) if r.actual and r.avg else 0,
                "Current Price":  round(r.actual or 0, 2),
                "Confidence":     "",
                "Signal":         "✓ In range" if in_range else "✗ Outside",
                "Score":          composite,
                "Grade":          _grade(composite),
                "Avg Error %":    round(mean_ape, 2),
                "Dir Accuracy %": round(dir_acc * 100, 1),
                "In Range %":     round(range_frac * 100, 1),
                "Matched Preds":  len(run_errors),
            })

        return score_rows

    # ------------------------------------------------------------------ #
    #  Excel: stock data                                                   #
    # ------------------------------------------------------------------ #

    def export_stock_data(self) -> str:
        """
        Write all OHLCV data to a fresh *STOCK_DATA_FILE*.

        Returns an absolute path on success, raises on failure.
        """
        with pd.ExcelWriter(STOCK_DATA_FILE, engine="openpyxl") as writer:
            for symbol in self._stocks:
                df_out = self.df_for_export(symbol)
                if df_out is not None:
                    df_out.to_excel(writer, sheet_name=symbol)
        return os.path.abspath(STOCK_DATA_FILE)

    def update_stock_data(self) -> str:
        """
        Append only new rows to *STOCK_DATA_FILE*, preserving history.

        Creates the file from scratch if it does not yet exist.
        Returns an absolute path on success, raises on failure.
        """
        if not os.path.exists(STOCK_DATA_FILE):
            return self.export_stock_data()

        from openpyxl import load_workbook

        wb = load_workbook(STOCK_DATA_FILE)

        for symbol in self._stocks:
            df_new = self.df_for_export(symbol)
            if df_new is None:
                continue

            if symbol in wb.sheetnames:
                existing = pd.read_excel(
                    STOCK_DATA_FILE, sheet_name=symbol,
                    index_col=0, engine="openpyxl",
                )
                existing.index = pd.to_datetime(existing.index).date
                last_date = existing.index.max()
                df_append = df_new[df_new.index > last_date]

                if df_append.empty:
                    self._cb("log", f"{symbol}: no new rows to append")
                    continue

                df_combined = pd.concat([existing, df_append])
                del wb[symbol]
                wb.save(STOCK_DATA_FILE)
            else:
                df_combined = df_new

            with pd.ExcelWriter(
                STOCK_DATA_FILE, engine="openpyxl",
                mode="a", if_sheet_exists="replace",
            ) as writer:
                df_combined.to_excel(writer, sheet_name=symbol)

            self._cb("log", f"{symbol}: stock data updated in {STOCK_DATA_FILE}")

        return os.path.abspath(STOCK_DATA_FILE)

    # ------------------------------------------------------------------ #
    #  Excel: predictions                                                  #
    # ------------------------------------------------------------------ #

    def export_predictions(self) -> str:
        """
        Write all current predictions to a fresh *PREDICTIONS_FILE*.

        Returns an absolute path on success, raises on failure.
        """
        with pd.ExcelWriter(PREDICTIONS_FILE, engine="openpyxl") as writer:
            for symbol in self._stocks:
                df_pred = self.pred_df_for_export(symbol)
                if df_pred is not None:
                    df_pred.to_excel(writer, sheet_name=symbol, index=False)
        return os.path.abspath(PREDICTIONS_FILE)

    def update_predictions(self) -> str:
        """
        Append new prediction rows to *PREDICTIONS_FILE*, preserving history.

        Creates the file from scratch if it does not yet exist.
        Returns an absolute path on success, raises on failure.
        """
        if not os.path.exists(PREDICTIONS_FILE):
            return self.export_predictions()

        for symbol in self._stocks:
            df_new = self.pred_df_for_export(symbol)
            if df_new is None:
                continue

            try:
                existing = pd.read_excel(
                    PREDICTIONS_FILE, sheet_name=symbol, engine="openpyxl"
                )
                df_combined = pd.concat([existing, df_new], ignore_index=True)
            except Exception:
                df_combined = df_new

            with pd.ExcelWriter(
                PREDICTIONS_FILE, engine="openpyxl",
                mode="a", if_sheet_exists="replace",
            ) as writer:
                df_combined.to_excel(writer, sheet_name=symbol, index=False)

            self._cb("log", f"{symbol}: predictions appended to {PREDICTIONS_FILE}")

        return os.path.abspath(PREDICTIONS_FILE)

    # ------------------------------------------------------------------ #
    #  Accuracy scoring                                                    #
    # ------------------------------------------------------------------ #

    def score_symbol_entry(self, symbol: str) -> ScoreResult:
        """
        Compute and store an accuracy score for *symbol*.
        """
        symbol = symbol.upper()
        data = self._stocks.get(symbol)
        if data is None:
            from stock_scorer import _insufficient_data_result
            return _insufficient_data_result(f"{symbol} is not tracked.")

        ph = data.get("pred_history", [])
        df = data.get("raw_df")
        cp = (data.get("prediction") or {}).get("current_price", 0.0)

        if df is None:
            from stock_scorer import _insufficient_data_result
            return _insufficient_data_result(f"{symbol} has no data.")

        result = score_symbol(ph, df, cp)
        data["accuracy_score"] = result
        self._cb("log",     f"{symbol} accuracy score: {result.score}/100 [{result.letter_grade}]")
        self._cb("refresh", symbol)
        return result

    def score_all_entries(self) -> dict:
        """
        Compute accuracy scores for every tracked symbol.
        """
        results = score_all(self._stocks)
        for symbol, result in results.items():
            self._stocks[symbol]["accuracy_score"] = result
            self._cb("log",     f"{symbol} accuracy score: {result.score}/100 [{result.letter_grade}]")
            self._cb("refresh", symbol)
        return results

    # ------------------------------------------------------------------ #
    #  Model persistence (xlsx)                                            #
    # ------------------------------------------------------------------ #

    def save_model_to_xlsx(self, symbol: str) -> None:
        """
        Persist the trained network weights, scaler params, and prediction
        history for *symbol* to MODELS_FILE.

        Sheet layout in stock_models.xlsx
        ----------------------------------
        ``{SYMBOL}``          – one row with the latest weights (overwritten
                                each save so the file stays compact).
        ``{SYMBOL}_history``  – append-only log of every archived prediction
                                (date, avg, best, worst).
        """
        data = self._stocks.get(symbol)
        if data is None:
            return
        system = data.get("system")
        if system is None:
            return

        m = system.model  # AdaptiveStockPredictor instance

        # ── weights row ────────────────────────────────────────────── #
        weights_row = {
            "Timestamp":    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "InputSize":    int(m.input_size),
            "HiddenSize":   int(m.hidden_size),
            "LR":           float(m.learning_rate),
            "W1":           json.dumps(m.W1.tolist()),
            "b1":           json.dumps(m.b1.tolist()),
            "W2":           json.dumps(m.W2.tolist()),
            "b2":           json.dumps(m.b2.tolist()),
            "ScalerParams": json.dumps(system.scaler_params),
            "FinalLoss":    float(m.losses[-1]) if m.losses else None,
        }
        df_weights = pd.DataFrame([weights_row])

        # ── prediction history rows ────────────────────────────────── #
        ph = data.get("pred_history", [])
        hist_rows = []
        for p in ph:
            dt = p["date"]
            hist_rows.append({
                "Date":  dt.strftime("%Y-%m-%d %H:%M:%S") if hasattr(dt, "strftime") else str(dt),
                "Avg":   round(float(p.get("avg",   0)), 4),
                "Best":  round(float(p.get("best",  0)), 4),
                "Worst": round(float(p.get("worst", 0)), 4),
            })
        df_hist = pd.DataFrame(hist_rows) if hist_rows else pd.DataFrame(
            columns=["Date", "Avg", "Best", "Worst"]
        )

        # ── write to xlsx ──────────────────────────────────────────── #
        try:
            mode = "a" if os.path.exists(MODELS_FILE) else "w"
            with pd.ExcelWriter(
                MODELS_FILE, engine="openpyxl",
                mode=mode, if_sheet_exists="replace",
            ) as writer:
                df_weights.to_excel(writer, sheet_name=symbol,              index=False)
                df_hist.to_excel(   writer, sheet_name=f"{symbol}_history", index=False)
            self._cb("log", f"{symbol}: model + history saved → {MODELS_FILE}")
        except Exception as exc:
            self._cb("log", f"Warning: could not save model for {symbol}: {exc}")

    def load_model_from_xlsx(self, symbol: str) -> bool:
        """
        Restore network weights, scaler params, and prediction history for
        *symbol* from MODELS_FILE.

        Returns True if the model was successfully restored, False otherwise.
        The caller is responsible for setting ``data["system"]`` before calling
        this — the weights are written directly into ``system.model``.
        """
        if not os.path.exists(MODELS_FILE):
            return False

        data = self._stocks.get(symbol)
        if data is None:
            return False
        system = data.get("system")
        if system is None:
            return False

        try:
            # ── restore weights ────────────────────────────────────── #
            df_w = pd.read_excel(MODELS_FILE, sheet_name=symbol, engine="openpyxl")
            if df_w.empty:
                return False

            row = df_w.iloc[-1]  # most recent saved state
            m = system.model

            m.W1 = np.array(json.loads(row["W1"]))
            m.b1 = np.array(json.loads(row["b1"]))
            m.W2 = np.array(json.loads(row["W2"]))
            m.b2 = np.array(json.loads(row["b2"]))

            sp = json.loads(row["ScalerParams"])
            # json round-trips int keys as strings — restore original keys
            system.scaler_params = {k: v for k, v in sp.items()}

            self._cb("log", f"{symbol}: weights restored from {MODELS_FILE} "
                            f"(saved {row['Timestamp']})")

            # ── restore prediction history ─────────────────────────── #
            try:
                df_h = pd.read_excel(
                    MODELS_FILE, sheet_name=f"{symbol}_history", engine="openpyxl"
                )
                ph = []
                for _, r in df_h.iterrows():
                    try:
                        dt = datetime.strptime(str(r["Date"]), "%Y-%m-%d %H:%M:%S")
                    except ValueError:
                        dt = datetime.strptime(str(r["Date"])[:10], "%Y-%m-%d")
                    ph.append({
                        "date":  dt,
                        "avg":   float(r["Avg"]),
                        "best":  float(r["Best"]),
                        "worst": float(r["Worst"]),
                    })
                data["pred_history"] = ph
                self._cb("log", f"{symbol}: {len(ph)} history point(s) restored")
            except Exception:
                pass  # history sheet missing — not fatal

            return True

        except Exception as exc:
            self._cb("log", f"{symbol}: no saved model found ({exc})")
            return False

    def _train_thread(self, symbol: str) -> None:
        try:
            data = self._stocks[symbol]
            system = StockTradingSystem(api_key="", lookback_window=data["lookback"])
            data["system"] = system  # set early so load_model_from_xlsx can access it

            end_date   = datetime.now().strftime("%Y-%m-%d")
            start_date = (datetime.now() - timedelta(days=180)).strftime("%Y-%m-%d")
            raw_df = system.api.fetch_stock_data(symbol, start_date, end_date)
            data["raw_df"] = raw_df

            # Try to resume from previously saved weights
            restored = self.load_model_from_xlsx(symbol)
            if restored:
                self._cb("status", f"Updating {symbol} (resumed from saved model)…")
                # Adapt existing weights to the latest data instead of full retrain
                system.train_model(symbol, epochs=max(20, data["epochs"] // 5))
            else:
                self._cb("status", f"Training {symbol} from scratch…")
                system.train_model(symbol, epochs=data["epochs"])

            data["status"] = "Trained"

            pred = system.predict_next_day(symbol, include_scenarios=True)
            data["prediction"] = pred
            data["status"] = "Ready"

            self.save_model_to_xlsx(symbol)

            self._cb("refresh", symbol)
            self._cb("chart",   symbol)
            self._cb("log",     f"✓ {symbol} {'updated' if restored else 'trained'} and predicted")
            self._cb("status",  f"Ready — {symbol} complete")
        except Exception as exc:
            err = str(exc)
            if symbol in self._stocks:
                self._stocks[symbol]["status"] = f"Error: {err[:40]}"
            self._cb("refresh", symbol)
            self._cb("log",     f"✗ {symbol}: {err}")

    def _predict_thread(self, symbol: str) -> None:
        try:
            data = self._stocks[symbol]
            if not data["system"]:
                self._cb("log", f"{symbol} not yet trained.")
                return
            self._cb("status", f"Predicting {symbol}…")

            pred = data["system"].predict_next_day(symbol, include_scenarios=True)
            self._archive_prediction(data)

            data["prediction"] = pred
            data["status"] = "Ready"

            self.save_model_to_xlsx(symbol)

            self._cb("refresh", symbol)
            self._cb("chart",   symbol)
            self._cb("log",     f"✓ {symbol} predictions refreshed")
            self._cb("status",  "Ready")
        except Exception as exc:
            self._cb("log", f"✗ {symbol} predict error: {exc}")

    def _update_thread(self, symbol: str) -> None:
        try:
            data = self._stocks[symbol]
            if not data["system"]:
                self._cb("log", f"{symbol} not yet trained.")
                return
            self._cb("status", f"Updating {symbol}…")

            data["system"].adaptive_update(symbol)
            self._archive_prediction(data)

            pred = data["system"].predict_next_day(symbol, include_scenarios=True)
            data["prediction"] = pred
            data["status"] = "Ready"

            self.save_model_to_xlsx(symbol)

            self._cb("refresh", symbol)
            self._cb("chart",   symbol)
            self._cb("log",     f"✓ {symbol} adaptively updated")
            self._cb("status",  "Ready")
        except Exception as exc:
            self._cb("log", f"✗ {symbol} update error: {exc}")

    # ------------------------------------------------------------------ #
    #  Private helpers                                                     #
    # ------------------------------------------------------------------ #

    def _archive_prediction(self, data: StockEntry) -> None:
        """Push the current prediction into pred_history before overwriting it."""
        old_pred = data.get("prediction")
        if old_pred and "scenarios" in old_pred:
            data["pred_history"].append({
                "date":  datetime.now(),
                "avg":   old_pred["scenarios"]["average_case"]["close"],
                "best":  old_pred["scenarios"]["best_case"]["close"],
                "worst": old_pred["scenarios"]["worst_case"]["close"],
            })
            ph = data.get("pred_history", [])
            df = data.get("raw_df")
            cp = (old_pred or {}).get("current_price", 0.0)
            if df is not None:
                data["accuracy_score"] = score_symbol(ph, df, cp)